from openpyxl import load_workbook

def buscar_encabezados(hoja):

    encabezados = {}

    for fila in hoja.iter_rows():

        for celda in fila:

            if celda.value is None:
                continue

            texto = str(celda.value).strip().lower()

            if texto == "articulo":
                encabezados["codigo"] = celda.column - 1

            elif texto == "descripcion":
                encabezados["descripcion"] = celda.column - 1

            elif texto == "precio/lista.":
                encabezados["precio"] = celda.column - 1

        if len(encabezados) == 3:

            encabezados["fila"] = fila[0].row
            return encabezados

    raise ValueError(
        "No se encontraron los encabezados obligatorios."
    )
def leer_proveedor(ruta_excel):

    libro = load_workbook(ruta_excel)
    hoja = libro.active

    encabezados = buscar_encabezados(hoja)

    print(f"Encabezados encontrados: {encabezados}")

    productos = []

    for fila in hoja.iter_rows(
        min_row=encabezados["fila"] + 1,
        values_only=True,
    ):
        
        codigo = fila[encabezados["codigo"]]
        descripcion = fila[encabezados["descripcion"]]
        precio = fila[encabezados["precio"]]

        if codigo is None:
            continue

        producto = {
            "codigo": int(codigo),
            "descripcion": descripcion,
            "precio": float(precio),
        }

        productos.append(producto)

    return productos